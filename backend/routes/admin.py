from flask import Blueprint, jsonify, request, send_file, session, make_response
from sqlalchemy import case, func

from ..extensions import db
from ..models import Book, User, BorrowRecord
from ..serializers import borrow_to_dict
from ..services.overdue import mark_overdue_records
from ..utils.auth import staff_required
from io import BytesIO
import datetime
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

admin_bp = Blueprint('admin', __name__)
ACTIVE_STATUSES = ('borrowed', 'overdue', 'pending_return')


@admin_bp.route('/dashboard', methods=['GET'])
@staff_required
def dashboard():
    mark_overdue_records()

    total_books = Book.query.count()
    total_users = User.query.count()
    active_borrows = BorrowRecord.query.filter(BorrowRecord.status.in_(ACTIVE_STATUSES)).count()
    overdue_count = BorrowRecord.query.filter_by(status='overdue').count()
    pending_return_count = BorrowRecord.query.filter_by(status='pending_return').count()
    available_copies = db.session.query(func.coalesce(func.sum(Book.available_quantity), 0)).scalar() or 0

    recent = (
        BorrowRecord.query.order_by(BorrowRecord.borrow_date.desc()).limit(8).all()
    )
    overdue_loans = (
        BorrowRecord.query.filter_by(status='overdue').order_by(BorrowRecord.due_date).limit(10).all()
    )
    pending_returns = (
        BorrowRecord.query.filter_by(status='pending_return').order_by(BorrowRecord.return_request_date.desc()).limit(10).all()
    )

    return jsonify({
        'stats': {
            'total_books': total_books,
            'total_users': total_users,
            'active_borrows': active_borrows,
            'overdue_count': overdue_count,
            'pending_return_count': pending_return_count,
            'available_copies': int(available_copies),
        },
        'recent_borrows': [borrow_to_dict(r, include_user=True) for r in recent],
        'overdue_loans': [borrow_to_dict(r, include_user=True) for r in overdue_loans],
        'pending_returns': [borrow_to_dict(r, include_user=True) for r in pending_returns],
    }), 200


def _collect_reports_data():
    mark_overdue_records()

    total_books = Book.query.count()
    total_users = User.query.count()
    total_borrows = BorrowRecord.query.count()
    active_borrows = BorrowRecord.query.filter(BorrowRecord.status.in_(ACTIVE_STATUSES)).count()
    overdue_count = BorrowRecord.query.filter_by(status='overdue').count()
    pending_return_count = BorrowRecord.query.filter_by(status='pending_return').count()
    unique_borrowers = db.session.query(func.count(func.distinct(BorrowRecord.user_id))).scalar() or 0
    active_borrowers = (
        db.session.query(func.count(func.distinct(BorrowRecord.user_id)))
        .filter(BorrowRecord.status.in_(ACTIVE_STATUSES))
        .scalar()
        or 0
    )
    available_copies = db.session.query(func.coalesce(func.sum(Book.available_quantity), 0)).scalar() or 0

    most_borrowed_books = (
        db.session.query(
            Book.id.label('book_id'),
            Book.title,
            Book.author,
            Book.genre,
            func.count(BorrowRecord.id).label('borrow_count'),
            func.sum(case((BorrowRecord.status.in_(ACTIVE_STATUSES), 1), else_=0)).label('active_count'),
            func.sum(case((BorrowRecord.status == 'overdue', 1), else_=0)).label('overdue_count'),
        )
        .join(BorrowRecord, BorrowRecord.book_id == Book.id)
        .group_by(Book.id)
        .order_by(func.count(BorrowRecord.id).desc(), Book.title.asc())
        .limit(10)
        .all()
    )

    overdue_reports = (
        BorrowRecord.query.filter_by(status='overdue').order_by(BorrowRecord.due_date).limit(20).all()
    )

    top_borrowers = (
        db.session.query(
            User.id,
            User.name,
            User.email,
            User.role,
            func.count(BorrowRecord.id).label('borrow_count'),
            func.sum(case((BorrowRecord.status.in_(ACTIVE_STATUSES), 1), else_=0)).label('active_count'),
            func.sum(case((BorrowRecord.status == 'overdue', 1), else_=0)).label('overdue_count'),
        )
        .join(BorrowRecord, BorrowRecord.user_id == User.id)
        .group_by(User.id)
        .order_by(func.count(BorrowRecord.id).desc(), User.name.asc())
        .limit(10)
        .all()
    )

    student_users = User.query.filter_by(role='student').count()
    staff_users = User.query.filter_by(role='admin').count()

    average_borrows_per_user = round(total_borrows / total_users, 2) if total_users else 0

    # Chart-friendly aggregates (last 30 days)
    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=29)

    borrows_by_date_q = (
        db.session.query(func.date(BorrowRecord.borrow_date).label('d'), func.count(BorrowRecord.id).label('c'))
        .filter(BorrowRecord.borrow_date >= start_date)
        .group_by(func.date(BorrowRecord.borrow_date))
        .all()
    )
    borrows_map = {r.d.isoformat(): int(r.c) for r in borrows_by_date_q}
    time_series = []
    for i in range(30):
        dt = start_date + datetime.timedelta(days=i)
        key = dt.isoformat()
        time_series.append({'date': key, 'borrows': borrows_map.get(key, 0)})

    overdue_by_date_q = (
        db.session.query(func.date(BorrowRecord.due_date).label('d'), func.count(BorrowRecord.id).label('c'))
        .filter(BorrowRecord.due_date >= start_date)
        .filter(BorrowRecord.status == 'overdue')
        .group_by(func.date(BorrowRecord.due_date))
        .all()
    )
    overdue_map = {r.d.isoformat(): int(r.c) for r in overdue_by_date_q}
    overdue_trend = []
    for i in range(30):
        dt = start_date + datetime.timedelta(days=i)
        key = dt.isoformat()
        overdue_trend.append({'date': key, 'overdue': overdue_map.get(key, 0)})

    borrow_by_genre_q = (
        db.session.query(Book.genre, func.count(BorrowRecord.id).label('c'))
        .join(BorrowRecord, BorrowRecord.book_id == Book.id)
        .group_by(Book.genre)
        .order_by(func.count(BorrowRecord.id).desc())
        .all()
    )
    borrow_by_genre = [{'genre': g or 'Unknown', 'count': int(c)} for g, c in borrow_by_genre_q]

    data = {
        'summary': {
            'total_books': total_books,
            'total_users': total_users,
            'total_borrows': total_borrows,
            'active_borrows': active_borrows,
            'overdue_count': overdue_count,
            'pending_return_count': pending_return_count,
            'unique_borrowers': int(unique_borrowers),
            'active_borrowers': int(active_borrowers),
            'available_copies': int(available_copies),
            'average_borrows_per_user': average_borrows_per_user,
        },
        'most_borrowed_books': [
            {
                'book_id': row.book_id,
                'title': row.title,
                'author': row.author,
                'genre': row.genre,
                'borrow_count': int(row.borrow_count or 0),
                'active_count': int(row.active_count or 0),
                'overdue_count': int(row.overdue_count or 0),
            }
            for row in most_borrowed_books
        ],
        'overdue_reports': [borrow_to_dict(record, include_user=True) for record in overdue_reports],
        'user_statistics': {
            'student_users': student_users,
            'staff_users': staff_users,
            'top_borrowers': [
                {
                    'user_id': row.id,
                    'name': row.name,
                    'email': row.email,
                    'role': row.role,
                    'borrow_count': int(row.borrow_count or 0),
                    'active_count': int(row.active_count or 0),
                    'overdue_count': int(row.overdue_count or 0),
                }
                for row in top_borrowers
            ],
        },
        'time_series': time_series,
        'overdue_trend': overdue_trend,
        'borrow_by_genre': borrow_by_genre,
    }

    return data


@admin_bp.route('/reports', methods=['GET'])
@staff_required
def reports():
    data = _collect_reports_data()
    return jsonify(data), 200


@admin_bp.route('/reports/export', methods=['GET'])
@staff_required
def export_reports():
    fmt = (request.args.get('format') or 'excel').strip().lower()
    data = _collect_reports_data()

    if fmt == 'excel' or fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summary'
        s = data.get('summary', {})
        ws.append(['Metric', 'Value'])
        for k, v in s.items():
            ws.append([k, v])

        # Most borrowed books
        mb = data.get('most_borrowed_books', [])
        ws2 = wb.create_sheet('MostBorrowed')
        ws2.append(['Title', 'Author', 'Genre', 'Borrows', 'Active', 'Overdue'])
        for b in mb:
            ws2.append([b.get('title'), b.get('author'), b.get('genre'), b.get('borrow_count'), b.get('active_count'), b.get('overdue_count')])

        # Top borrowers
        tb = data.get('user_statistics', {}).get('top_borrowers', [])
        ws3 = wb.create_sheet('TopBorrowers')
        ws3.append(['Name', 'Email', 'Role', 'Borrows', 'Active', 'Overdue'])
        for u in tb:
            ws3.append([u.get('name'), u.get('email'), u.get('role'), u.get('borrow_count'), u.get('active_count'), u.get('overdue_count')])

        # All Circulation Records
        ws4 = wb.create_sheet('Circulation')
        ws4.append(['Borrower Name', 'Borrower Email', 'Book Title', 'Due Date', 'Status'])
        all_borrows = BorrowRecord.query.order_by(BorrowRecord.borrow_date.desc()).all()
        for record in all_borrows:
            due_date_str = record.due_date.strftime('%Y-%m-%d %I:%M %p') if record.due_date else '—'
            ws4.append([
                record.user.name if record.user else 'Unknown',
                record.user.email if record.user else '',
                record.book.title if record.book else 'Unknown Book',
                due_date_str,
                record.status
            ])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"library-reports-{datetime.date.today().isoformat()}.xlsx"
        return send_file(bio, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # PDF export
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        spaceAfter=15,
        textColor=colors.HexColor('#1E293B')
    )

    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        spaceBefore=14,
        spaceAfter=8,
        textColor=colors.HexColor('#0F172A')
    )

    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )

    cell_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    story = []

    # Title
    story.append(Paragraph(f"Library Reports - {datetime.date.today().strftime('%B %d, %Y')}", title_style))
    story.append(Spacer(1, 10))

    # System Summary Table
    story.append(Paragraph("System Summary", h2_style))
    summary_data = [
        [Paragraph("Metric", cell_header_style), Paragraph("Value", cell_header_style)]
    ]

    s = data.get('summary', {})
    metric_labels = {
        'total_books': 'Total Books',
        'total_users': 'Total Users',
        'total_borrows': 'Total Borrows',
        'active_borrows': 'Active Borrows',
        'overdue_count': 'Overdue Borrows',
        'pending_return_count': 'Pending Returns',
        'unique_borrowers': 'Unique Borrowers',
        'active_borrowers': 'Active Borrowers',
        'available_copies': 'Available Copies',
        'average_borrows_per_user': 'Avg Borrows/User'
    }
    for k, v in s.items():
        label = metric_labels.get(k, k.replace('_', ' ').title())
        summary_data.append([Paragraph(label, cell_style), Paragraph(str(v), cell_style)])

    summary_table = Table(summary_data, colWidths=[240, 300], hAlign='LEFT')
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#1E293B')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 10))

    # Most Borrowed Books Table
    story.append(Paragraph("Most Borrowed Books", h2_style))
    mb_data = [
        [
            Paragraph("Title", cell_header_style),
            Paragraph("Author", cell_header_style),
            Paragraph("Genre", cell_header_style),
            Paragraph("Borrows", cell_header_style),
            Paragraph("Active", cell_header_style),
            Paragraph("Overdue", cell_header_style)
        ]
    ]
    for b in data.get('most_borrowed_books', []):
        mb_data.append([
            Paragraph(b.get('title', ''), cell_style),
            Paragraph(b.get('author', ''), cell_style),
            Paragraph(b.get('genre', ''), cell_style),
            Paragraph(str(b.get('borrow_count', 0)), cell_style),
            Paragraph(str(b.get('active_count', 0)), cell_style),
            Paragraph(str(b.get('overdue_count', 0)), cell_style)
        ])
    mb_table = Table(mb_data, colWidths=[160, 100, 80, 70, 65, 65], hAlign='LEFT')
    mb_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(mb_table)
    story.append(Spacer(1, 10))

    # Top Borrowers Table
    story.append(Paragraph("Top Borrowers", h2_style))
    tb_data = [
        [
            Paragraph("Name", cell_header_style),
            Paragraph("Email", cell_header_style),
            Paragraph("Role", cell_header_style),
            Paragraph("Borrows", cell_header_style),
            Paragraph("Active", cell_header_style),
            Paragraph("Overdue", cell_header_style)
        ]
    ]
    for u in data.get('user_statistics', {}).get('top_borrowers', []):
        tb_data.append([
            Paragraph(u.get('name', ''), cell_style),
            Paragraph(u.get('email', ''), cell_style),
            Paragraph(u.get('role', ''), cell_style),
            Paragraph(str(u.get('borrow_count', 0)), cell_style),
            Paragraph(str(u.get('active_count', 0)), cell_style),
            Paragraph(str(u.get('overdue_count', 0)), cell_style)
        ])
    tb_table = Table(tb_data, colWidths=[100, 160, 60, 75, 75, 70], hAlign='LEFT')
    tb_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(tb_table)
    story.append(Spacer(1, 10))

    # All Circulation Records Table
    story.append(Paragraph("All Circulation Records", h2_style))
    circ_data = [
        [
            Paragraph("Borrower", cell_header_style),
            Paragraph("Borrowed Book", cell_header_style),
            Paragraph("Due Date", cell_header_style),
            Paragraph("Status", cell_header_style)
        ]
    ]

    all_borrows = BorrowRecord.query.order_by(BorrowRecord.borrow_date.desc()).all()
    for record in all_borrows:
        borrower_info = f"<b>{record.user.name if record.user else 'Unknown'}</b><br/><font color='#64748B'>{record.user.email if record.user else ''}</font>"
        book_title = record.book.title if record.book else 'Unknown Book'
        due_date_str = record.due_date.strftime('%Y-%m-%d %I:%M %p') if record.due_date else '—'
        status_str = record.status
        circ_data.append([
            Paragraph(borrower_info, cell_style),
            Paragraph(book_title, cell_style),
            Paragraph(due_date_str, cell_style),
            Paragraph(status_str, cell_style)
        ])

    circ_table = Table(circ_data, colWidths=[180, 180, 110, 70], hAlign='LEFT')
    circ_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(circ_table)

    doc.build(story)
    bio.seek(0)
    filename = f"library-reports-{datetime.date.today().isoformat()}.pdf"
    return send_file(bio, download_name=filename, as_attachment=True, mimetype='application/pdf')


def user_to_dict(user):
    return {
        'id': user.id,
        'student_id': user.student_id,
        'name': user.name,
        'email': user.email,
        'role': user.role,
        'is_active': user.is_active,
        'created_at': user.created_at.isoformat() if user.created_at else None
    }


@admin_bp.route('/users', methods=['GET'])
@staff_required
def get_users():
    users = User.query.all()
    return jsonify({'users': [user_to_dict(u) for u in users]}), 200


@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@staff_required
def update_user(user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json() or {}
    
    if 'name' in data:
        user.name = data['name']
    if 'email' in data:
        existing = User.query.filter_by(email=data['email']).first()
        if existing and existing.id != user.id:
            return jsonify({'error': 'Email is already taken'}), 400
        user.email = data['email']
    if 'student_id' in data:
        existing = User.query.filter_by(student_id=data['student_id']).first()
        if existing and existing.id != user.id:
            return jsonify({'error': 'Student ID is already taken'}), 400
        user.student_id = data['student_id']
    if 'role' in data:
        if data['role'] not in ('admin', 'student'):
            return jsonify({'error': 'Invalid role'}), 400
        user.role = data['role']
    if 'is_active' in data:
        user.is_active = bool(data['is_active'])
        if user.is_active:
            user.failed_login_attempts = 0
        
    db.session.commit()
    return jsonify({'message': 'User updated successfully', 'user': user_to_dict(user)}), 200


@admin_bp.route('/users/<int:user_id>', methods=['DELETE'])
@staff_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    current_user_id = session.get('user_id')
    if current_user_id == user.id:
        return jsonify({'error': 'You cannot delete yourself'}), 400
        
    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'User deleted successfully'}), 200


@admin_bp.route('/backup', methods=['GET'])
@staff_required
def backup_database():
    try:
        conn = db.engine.raw_connection()
        cursor = conn.cursor()
        
        cursor.execute("SHOW TABLES")
        tables = [row[0] for row in cursor.fetchall()]
        
        sql_dump = []
        sql_dump.append("-- SmartLib Database Backup")
        sql_dump.append(f"-- Generated on: {datetime.datetime.utcnow().isoformat()} UTC\n")
        sql_dump.append("CREATE DATABASE IF NOT EXISTS smartlib;")
        sql_dump.append("USE smartlib;\n")
        
        for table in tables:
            cursor.execute(f"SHOW CREATE TABLE {table}")
            create_stmt = cursor.fetchone()[1]
            sql_dump.append(f"DROP TABLE IF EXISTS `{table}`;")
            sql_dump.append(f"{create_stmt};\n")
            
            cursor.execute(f"SELECT * FROM `{table}`")
            rows = cursor.fetchall()
            if rows:
                cursor.execute(f"DESCRIBE `{table}`")
                columns = [col[0] for col in cursor.fetchall()]
                col_names = ", ".join(f"`{c}`" for c in columns)
                
                for row in rows:
                    values = []
                    for val in row:
                        if val is None:
                            values.append("NULL")
                        elif isinstance(val, (int, float)):
                            values.append(str(val))
                        elif isinstance(val, (datetime.datetime, datetime.date)):
                            values.append(f"'{val.isoformat()}'")
                        elif isinstance(val, bytes):
                            values.append(f"X'{val.hex()}'")
                        else:
                            escaped = str(val).replace("'", "''").replace("\\", "\\\\")
                            values.append(f"'{escaped}'")
                    val_str = ", ".join(values)
                    sql_dump.append(f"INSERT INTO `{table}` ({col_names}) VALUES ({val_str});")
                sql_dump.append("")
                
        cursor.close()
        conn.close()
        
        dump_content = "\n".join(sql_dump)
        response = make_response(dump_content)
        response.headers["Content-Disposition"] = "attachment; filename=smartlib_backup.sql"
        response.headers["Content-Type"] = "application/sql"
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500
